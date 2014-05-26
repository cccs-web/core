"""
Analysis utillities for the CV spreadsheets.
"""
from openpyxl import load_workbook

from importers.cvs_globals import *


def distinct_column_values(sheet_name, column_index):
    """
    Return a dict of the relevant column values mapping to the count of those values in the columns for all cvs.
    """
    # Just a utility so I don't worry about fetching the heading.

    values = dict()
    for cv in xlsx_cvs:
        try:
            wb = load_workbook(filename=cv)
        except:
            print("Could not load '{0}'".format(cv))
            continue
        ws = wb.get_sheet_by_name(sheet_name)
        for cell in ws.columns[column_index]:
            value = cell.value
            if value is not None:
                values[value] = values.get(value, 0) + 1
    return values


def get_rows(sheet_name, row_index):
    result = list()
    for cv in xlsx_cvs:
        try:
            wb = load_workbook(filename=cv)
        except:
            print("Could not load '{0}'".format(cv))
            continue
        ws = wb.get_sheet_by_name(sheet_name)
        result.append([c.value for c in ws.rows[row_index]])
    return result