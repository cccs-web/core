"""
Load in initial data using the source excel spreadsheets
"""
import datetime
import re
import calendar

from openpyxl import load_workbook

from mezzanine.utils.urls import slugify

import projects.models as pm
import cvs.models as cm
from importers.cvs_globals import *


# I thought it would be nice to have all the sheet names here for clarity but it isn't. I'm gradually
# phasing them out for literals because it is (for once) clearer.
def load_all():
    load_countries('.country-calc, CCCS')
    load_categorizations(get_theme_dict('.theme-calc, CCCS'), pm.CCCSTheme, pm.CCCSSubTheme, 'theme')
    load_categorizations(get_theme_dict('.theme-calc, IFC'), pm.IFCTheme, pm.IFCSubTheme, 'theme')
    load_ifc_sectors('.sector-calc, IFC')
    load_categorizations(get_cccs_sector_dict('.sector-calc, CCCS'), pm.CCCSSector, pm.CCCSSubSector, 'sector')
    load_projects()
    load_cvs()


def load_countries(sheet_name):
    country_dicts = get_country_dicts(sheet_name)
    for info in country_dicts:
        country, _ = pm.Country.objects.get_or_create(name=info['name'])
        for (k, v) in info.iteritems():
            setattr(country, k, v)
        country.save()


def get_country_dicts(sheet_name):
    """
    Use the gold cv to obtain the list of countries and related data
    Go for the '.country-calc, CCCS' sheet and collect the data from row 4 to the end.
    """
    wb = load_workbook(filename=xlsx_cv)
    ws = wb.get_sheet_by_name(sheet_name)

    countries = list()
    for row in ws.rows[3:]:  # ignore first three rows
        if row[0].value is None:
            continue  # ignore null entries
        countries.append({
            'name': row[0].value,
            'iso_english_name': row[1].value,
            'fips': row[2].value,
            'iso_numeric': row[3].value,
            'iso_3166': row[4].value,
            'iso': row[5].value,
            'notes': row[6].value})
    return countries


def load_categorizations(info, theme_model, sub_theme_model, super_field_name):
    for (theme_name, sub_theme_names) in info.iteritems():
        for sub_theme_name in sub_theme_names:
            _load_categorizations(theme_model, theme_name, sub_theme_model, sub_theme_name, super_field_name)


def get_theme_dict(sheet_name):
    """
    Use the gold cv to obtain all the cccs themes and sub themes
    """
    wb = load_workbook(filename=xlsx_cv)
    ws = wb.get_sheet_by_name(sheet_name)

    themes = dict()
    for row in ws.rows[1:]:  # ignore heading row
        theme_name = row[3].value
        sub_theme_name = row[4].value
        if theme_name is None:
            continue
        if theme_name not in themes:
            themes[theme_name] = set()
        themes[theme_name].add(sub_theme_name)
    return themes


def load_ifc_sectors(sheet_name):
    sector_names = get_ifc_sector_names(sheet_name)
    for sector_name in sector_names:
        sector, created = pm.IFCSector.objects.get_or_create(name=sector_name)
        if created:
            sector.save()


def get_ifc_sector_names(sheet_name):
    """
    Use the gold cv to obtain all the IFC sectors
    """
    wb = load_workbook(filename=xlsx_cv)
    ws = wb.get_sheet_by_name(sheet_name)
    sector_names = list()
    for row in ws.rows[2:]:  # ignore first two rows
        sector_name = row[1].value
        if sector_name is None:
            continue
        sector_names.append(sector_name)
    return sector_names


def get_cccs_sector_dict(sheet_name):
    """
    Use the gold cv to obtain all the cccs themes and sub themes
    """
    wb = load_workbook(filename=xlsx_cv)
    ws = wb.get_sheet_by_name(sheet_name)

    sectors = dict()
    for row in ws.rows[1:]:  # ignore heading row
        sector_name = row[1].value
        sub_sector_name = row[2].value
        if sector_name is None:
            continue
        if sector_name not in sectors:
            sectors[sector_name] = set()
        sectors[sector_name].add(sub_sector_name)
    return sectors


def load_projects():
    projects = get_project_dict()
    for project_name, project_info in projects.iteritems():
        project, _ = pm.Project.objects.get_or_create(title=project_name)
        project.title = project_name

        def gp(key):
            return project_info.get(key)

        for (attname, key) in (('date_range', 'Date Range'),
                               ('loan_or_grant', 'Loan/TA/Grant No'),
                               ('features', 'Main Project Features'),
                               ('region', 'Region'),
                               ('position', 'Position'),
                               ('client_end', 'Sponsor / End Client'),
                               ('client_contract', 'Contracted Through / Direct Client'),
                               ('client_beneficiary', 'Beneficiary Client'),
                               ('contract', 'Contract No.'),
                               ('activities', 'Activities Performed'),
                               ('references', 'References')):
            setattr(project, attname, gp(key))

        # Multiple countries are allowed but only one from import
        country = get_country(gp('Country'))
        if country:
            project.countries.add(country)
        else:
            print(u"No country for \"{0}\": '{1}' does not exist".format(project_name, gp('Country')))

        # Break Services On/Off-site into flags and have a go at importing them
        services = project_info.get('Services On/Off-site', None)
        if services is not None:
            project.service_on_site = 'On' in services
            project.service_off_site = 'Off' in services
            project.service_on_site = 'remote' in services

        for attname, args in (
                ('cccs_subthemes', (pm.CCCSTheme, gp('Thematic Issues -GENERAL'),
                                    pm.CCCSSubTheme, gp('Sub-Themes -GENERAL'),
                                    'theme')),
                ('cccs_subsectors', (pm.CCCSSector, gp('Sector -GENERAL'),
                                     pm.CCCSSubSector, gp('Sub-sector -GENERAL'),
                                     'sector')),
                ('ifc_subthemes', (pm.IFCTheme, gp('Thematic Issues -IFC'),
                                   pm.IFCSubTheme, gp('Sub-Themes -IFC'),
                                   'theme', False))):
            for categorization in _load_categorizations(*args):
                getattr(project, attname).add(categorization)

        ifc_sector_names = get_names_from_string(gp('Sector -IFC'))
        for ifc_sector_name in ifc_sector_names:
            ifc_sector, created = pm.IFCSector.objects.get_or_create(name=ifc_sector_name)
            if created:
                ifc_sector.save()
            project.ifc_sectors.add(ifc_sector)

        project.save()


def _load_categorizations(super_class, super_values, sub_class, sub_values, super_field, add_unspecified=True):
    """
    Get or create the categorizations returning sub instances
    """
    if super_values:
        super_names = get_names_from_string(super_values)
    else:
        super_names = []
    if sub_values:
        sub_names = get_names_from_string(sub_values)
    else:
        sub_names = []

    if len(super_names) != len(sub_names):
        print("Mismatched theme/subtheme list lengths in cv - using shortest")
    if add_unspecified and not(super_names and sub_names):
        super_names = sub_names = ['Unspecified']

    sub_objects = list()
    for super_name, sub_name in zip(super_names, sub_names):
        super_obj, created = super_class.objects.get_or_create(name=super_name)
        if created:
            super_obj.save()

        sub_kwargs = {super_field: super_obj, 'name': sub_name}
        sub_obj, created = sub_class.objects.get_or_create(**sub_kwargs)
        if created:
            sub_obj.save()
        sub_objects.append(sub_obj)

    return sub_objects


def get_names_from_string(s):
    if s is None:
        return []
    else:
        return [name.strip() for name in s.split(';') if name]


def get_project_dict():
    projects = dict()
    for cv in xlsx_cvs:
        try:
            wb = load_workbook(filename=cv)
        except TypeError:
            print("Unable to open {0}".format(cv))
            continue
        _get_project_dicts_from_wb(wb, projects)
    return projects


def _get_project_dicts_from_wb(wb, projects=None):
    """
    Get project dicts from ws, adding them into projects if supplied.
    """
    if projects is None:
        projects = dict()

    project_dicts = _get_all_row_dicts(wb, 'PROJECT', 1)
    for project_info in project_dicts:
        project_name = project_info['Assignment or Project Name']
        if project_name in projects:
            # Add any values we don't already have
            existing_info = projects[project_name]
            for k, val in project_info.items():
                if k not in existing_info:
                    existing_info[k] = val
                elif val != existing_info[k]:
                    print(u"Project {0} [{1}] candidates: '{2}' vs '{3}'".format(
                        project_name,
                        k,
                        val,
                        existing_info[k]))
                    # Use longest or biggest
                    try:
                        if len(val) > len(existing_info[k]):
                            existing_info[k] = val
                    except TypeError:
                        try:
                            if val > existing_info[k]:
                                existing_info[k] = val
                        except TypeError:
                            pass
        else:
            projects[project_name] = project_info
    return projects


def _get_edutraining_dicts_from_wb(wb):
    """
    Get edutraining as a list of lists and access by index (it is a mess).
    """
    ws = wb.get_sheet_by_name('EDU _ TRAINING')
    return [[cell.value for cell in row] for row in ws.rows[2:]]


def _get_all_row_dicts(wb, sheet_name, test_row_index):
    ws = wb.get_sheet_by_name(sheet_name)
    headings = _get_column_headings(ws.rows[0])
    employment_list = list()
    for row in ws.rows[1:]:
        test_value = row[test_row_index].value
        if test_value is not None:
            info = dict()
            for k, c in zip(headings, row):
                value = c.value
                if value is None:
                    continue
                info[k] = value
            employment_list.append(info)
    return employment_list


def _get_column_headings(heading_row):
    headings = list()
    for heading_cell in heading_row:
        if heading_cell.value is None:
            break
        headings.append(canonical(heading_cell.value.strip()))
    return headings


def load_cvs():
    cv_dicts = get_cv_dicts()
    for cv_dict in cv_dicts:
        try:
            cv = cm.CV.objects.get(slug=cv_dict['slug'])
        except cm.CV.DoesNotExist:
            cv = cm.CV()
        cv.slug = cv_dict['slug']

        setup_user(cv_dict, cv)

        for (attname, key) in (('middle_names', 'Given Middle Name'),
                               ('alternate_names', 'Alternate Names Used'),
                               ('street', 'Mailing Address'),
                               ('city', 'City'),
                               ('state', 'State/Province'),
                               ('zip', 'ZIP/Postal Code'),
                               ('telephone', 'Telephone'),):
            setattr(cv, attname, cv_dict.get(key))

        # do the country fields together
        for (attname, key) in (('country', 'Country'),
                               ('citizenship', 'Citizenship'),
                               ('birth_country', 'Country of Birth')):
            country = get_country(cv_dict.get(key))
            if country:
                setattr(cv, attname, country)
            else:
                print("No country:", cv_dict['fname'], cv_dict.get(key))

        dob = cv_dict.get('Date of Birth')
        if type(dob) == datetime.datetime:
            cv.dob = dob

        # gender and marital status use the first letter of the relevant word in the field
        for (attname, key) in (('gender', 'Sex'),
                               ('marital_status', 'Marital Status')):
            value = cv_dict.get(key)
            if value:
                setattr(cv, attname, value[0].upper())

        cv.save()  # This must be done here so there is an entity to link related objects to

        for project_info in cv_dict['projects']:
            project_name = project_info['Assignment or Project Name']
            project = pm.Project.objects.get(title=project_name)
            cv_project, _ = cm.CVProject.objects.get_or_create(cv=cv, project=project)
            cv_project.project = project
            cv_project.cv = cv
            for (attname, key) in (
                    ('position', 'Position'),
                    ('person_months', 'Person-months'),
                    ('activities', 'Activities Performed'),
                    ('references', 'References')):
                setattr(cv_project, attname, project_info.get(key))
            cv_project.save()

        for row in cv_dict['edutraining']:
            _add_cv_education(row[0:5], cv)
            _add_cv_training(row[5:8], cv)
            _add_cv_membership(row[8:11], cv)
            _add_cv_languages(row[11:15], cv)

        _add_cv_employment(cv_dict['employment'], cv)
        _add_cv_publication(cv_dict['publications'], cv)


def _set_values(obj, value_set):
    for (attname, value) in value_set:
        if value:
            setattr(obj, attname, value)


def _set_dates(obj, period):
    for (attname, dt) in zip(('from_date', 'to_date'), _build_period(period)):
        if dt:
            setattr(obj, attname, dt)


def _add_cv_education(row, cv):
    """
    Add cv education entries
    """
    (institution, years, degree, majors, minors) = row
    if majors or minors:
        subject = _build_education_subject(majors, minors)
        education, _ = cm.CVEducation.objects.get_or_create(cv=cv, subject=subject)
        _set_values(education, (
            ('cv', cv),
            ('subject', subject),
            ('institution', institution),
            ('qualification', degree)))
        _set_dates(education, years)
        education.save()


def _add_cv_training(row, cv):
    (institution, years, subject) = row
    if subject:
        training, _ = cm.CVTraining.objects.get_or_create(cv=cv, subject=subject, institution=institution)
        _set_values(training, (
            ('cv', cv),
            ('subject', subject),
            ('institution', institution)))
        _set_dates(training, years)
        training.save()


def _add_cv_membership(row, cv):
    (organization, years, role) = row
    if organization:
        membership, _ = cm.CVMembership.objects.get_or_create(cv=cv, organization=organization)
        _set_values(membership, (
            ('cv', cv),
            ('organization', organization),
            ('role', role)))
        _set_dates(membership, years)
        membership.save()


def _add_cv_languages(row, cv):
    (dialect, reading, speaking, writing) = row
    if dialect:
        language, created = cm.Language.objects.get_or_create(name=dialect)
        cvlanguage, _ = cm.CVLanguage.objects.get_or_create(cv=cv, language=language)
        _set_values(cvlanguage, (
            ('cv', cv),
            ('language', language),
            ('reading', reading),
            ('speaking', speaking),
            ('writing', writing)))
        cvlanguage.save()


def _add_cv_employment(employments, cv):
    for info in employments:
        info['from_date'], info['to_date'] = _build_period(info['Dates'], True)
        employment, created = cm.CVEmployment.objects.get_or_create(
            cv=cv,
            employer=info['Employer'],
            from_date=info['from_date'],
            to_date=info['to_date'])
        _set_values(employment, (
            ('cv', cv),
            ('from_date', info['from_date']),
            ('to_date', info['to_date']),
            ('employer', info['Employer']),
            ('location', info.get('Location')),
            ('position', info.get('Position')),
            ('accomplishments', info.get('Duties and Accomplishments')),
            ('references', info.get('Reference(s)'))))
        employment.save()


def _add_cv_publication(publications, cv):
    for info in publications:
        title = info['Title']
        publication, created = cm.CVPublication.objects.get_or_create(
            cv=cv,
            title=title)
        _set_values(publication, (
            ('cv', cv),
            ('publication_date', _build_date(info.get('Date'))),
            ('publication_type', info.get('Type')),
            ('author', info.get('Editors/Authors')),
            ('title', title),
            ('distribution', info.get('Distribution')),
            ('identifier', info.get('ID#'))))
        try:
            publication.save()
        except:
            import pdb; pdb.set_trace()


def _build_education_subject(majors, minors):
    if majors and minors:
        return u"{0} with {1}".format(majors, minors)
    elif majors:
        return majors
    elif minors:
        return minors
    else:
        raise Exception("majors and minors not specified.")

short_month_names = '|'.join(calendar.month_abbr)[1:]
long_month_names = '|'.join(calendar.month_name)[1:]
month_re = '({0}|{1})'.format(short_month_names, long_month_names)


def _build_date(date_str):
    """
    date_str could be a use mm/dd/yyyy, mm/dd/yy or funny stuff such as:
    "2007, Sept 10-12"
    We need to try for the most likely meaningful date.
    """
    if date_str is None:
        return None

    if isinstance(date_str, int):
        return datetime.datetime(year=int(date_str), month=1, day=1)

    if isinstance(date_str, datetime.datetime):
        return date_str

    year = month = day = None

    # Try simple US formats (allow any separator character)
    match = re.match(r'\b(\d+).(\d+).(\d+)\b', date_str)
    if match:
        (month, day, year) = [int(i) for i in match.groups()]
        if year < 100:  # Fix to four digits
            if year > 14:
                year += 1900
            else:
                year += 2000
    else:
        # Try extracting year, month and day one by one
        match = re.search(r'(\d{4})', date_str)  # 2 digit year is too hard here
        if match is None:
            return None

        year_str = match.groups()[0]
        date_str = date_str.replace(year_str, '')
        year = int(year_str)

        match = re.search(month_re, date_str)
        if match:
            month_str = match.groups()[0]
            date_str = date_str.replace(month_str, '')
            month = int(month_to_month_number(month_str))

        match = re.search(r'(\d+)', date_str)
        if match:
            day_str = match.groups()[0]
            day = int(day_str)

    if year is None:
        return None

    if month is None or (month < 0 or month > 13):
        month = 1

    first_day, last_day = calendar.monthrange(year, month)
    if day < first_day or day > last_day:
        day = 1

    return datetime.datetime(year=year, month=month, day=day)


def _build_period(period, use_from_for_to=False):
    """
    period is often just one year e.g. 1966, often a hyphenated range e.g. "1967-1971",
    sometimes a list of years,  and sometmes some sort of date.
    Years are almost always four digit.
    """
    if period is None:
        return None, None
    if isinstance(period, int):
        dt = datetime.datetime(year=int(period), month=1, day=1)
        return dt, dt if use_from_for_to else None

    if use_from_for_to:
        # Change it back if there is the word 'present' or 'current' in the period
        use_from_for_to = not re.search(r'present|current', period, re.IGNORECASE)

    # The trick I've used here is to break out a list of possible years and months and then use those to build the
    # likely from and to dates rather than faffing about with complicated regular expressions.
    years = re.findall(r'(\d{4})', period)
    months = re.findall(month_re, period, re.IGNORECASE)

    if not years:
        return None, None

    from_year = years[0]
    from_month = 'Jan'
    if months:
        from_month = months[0]

    to_year = None
    to_month = 'Dec'
    if years[1:]:
        to_year = years[1]
    elif use_from_for_to:
        to_year = from_year
    if months[1:]:
        to_month = months[1]

    from_dt = datetime.datetime(year=int(from_year),
                                month=month_to_month_number(from_month),
                                day=1)
    if to_year:
        to_year = int(to_year)
        to_month_num = month_to_month_number(to_month)
        to_dt = datetime.datetime(year=to_year,
                                  month=to_month_num,
                                  day=calendar.monthrange(to_year, to_month_num)[1])
    else:
        to_dt = None

    return from_dt, to_dt


def month_to_month_number(month_str):
    try:
        return datetime.datetime.strptime(month_str, '%b').month
    except ValueError:
        try:
            return datetime.datetime.strptime(month_str, '%B').month
        except ValueError:  # improve message
            raise ValueError("time data {0} is not a month name or abbreviation".format(month_str))


def _get_country_by_field(s, field_name):
    kwargs = {field_name: s}
    try:
        return pm.Country.objects.get(**kwargs)
    except pm.Country.DoesNotExist:
        return None


def _get_country_by_name(s):
    s = canonical(s)
    return _get_country_by_field(s, 'name')


def _get_country_by_iso(s):
    # strip out any periods
    s = ''.join(s.split('.'))
    return _get_country_by_field(s, 'iso_3166')

country_names = pm.Country.objects.all().values_list('name', flat=True)


def _guess_country(s):
    try:
        country_name = next((country_name for country_name in country_names if country_name in s))
        return pm.Country.objects.get(name=country_name)
    except StopIteration:
        return None


def get_country(s):
    """
    Try to return a country matching s
    """
    if s:
        for f in (_get_country_by_name,
                  _get_country_by_iso,
                  _guess_country):
            country = f(s)
            if country:
                return country


def setup_user(cv_dict, cv):
    """
    Use existing user if available on cv object. If no user:
    Find an existing matching user or create one as necessary.
    """
    try:
        return cv.user
    except cm.User.DoesNotExist:
        username = None
        for username in gen_username(cv_dict['Given First Name'], cv_dict.get('Surname')):
            try:
                if cm.CV.objects.get(user__username=username):
                    continue
            except cm.CV.DoesNotExist:
                break

        user, created = cm.User.objects.get_or_create(username=username)

        if created:
            user.username = username
            user.set_password(u"{0}42why".format(username))
            user.first_name = cv_dict['Given First Name']
            user.last_name = cv_dict['Surname']
            if cv_dict.get('Email'):
                user.email = cv_dict['Email']
            user.save()

        cv.user = user

        return user


def gen_username(first_name, last_name=None, max_len=30):
    """
    Generate username candidates
    'Jo Ho' -> ['jo', 'joh', 'joho', 'jo1', 'jo2'...]
    """
    # TODO: This may duplicate integer suffixed usernames when cropping to max_len because the duplicate
    # may be ten earlier e.g:
    # [n for n in gen_username('a', '', max_len=2)] ->
    # ['a', 'a1', 'a2', 'a3', 'a4', 'a5', 'a6', 'a7', 'a8', 'a9', 'a1']

    if last_name is None:
        last_name = ''
    if max_len < 1:
        raise Exception('gen_username: max_len must be greater than zero')
    if len(first_name) < 1:
        raise Exception('gen_username: first_name length must be greater than zero')

    previous_name_stash = [None]  # Controls when to stop

    def build(first, second, previous_name_stash):
        len_first, len_second = len(first), len(second)
        reduction = max(0, (len_first + len_second) - max_len)
        first = first[0:max(1, len_first - reduction)]
        reduction = max(0, reduction - (len_first - 1))
        second = second[0:max(0, len_second - reduction)]

        new_name = first + second
        if new_name == previous_name_stash[0]:
            raise StopIteration
        previous_name_stash[0] = new_name
        return new_name

    first_name = first_name.lower()
    last_name = last_name.lower()

    yield build(first_name, '', previous_name_stash)
    for i in range(len(last_name)):
        yield build(first_name, last_name[0:i+1], previous_name_stash)
    i = 0
    while True:
        i += 1
        yield build(first_name, str(i), previous_name_stash)


def get_cv_dicts():
    cv_dicts = list()
    for cv in xlsx_cvs:
        try:
            wb = load_workbook(filename=cv)
        except TypeError:
            print("Unable to open {0}".format(cv))
            continue
        ws = wb.get_sheet_by_name('BIODATA')
        cv_dict = {canonical(row[0].value.strip()): row[1].value
                   for row in ws.rows if row[0].value is not None}
        # slug is not perfect (assumes no duplicate names)
        # but is needed here so we can update cvs effectively
        if 'Given First Name' in cv_dict:
            cv_dict['fname'] = cv
            cv_dict['slug'] = get_slug(cv_dict)
            cv_dict['projects'] = _get_all_row_dicts(wb, 'PROJECT', 1)
            cv_dict['edutraining'] = _get_edutraining_dicts_from_wb(wb)
            cv_dict['employment'] = _get_all_row_dicts(wb, 'EMPLOYMENT', 1)
            cv_dict['publications'] = _get_all_row_dicts(wb, 'PUBLICATIONS', 3)
            cv_dicts.append(cv_dict)
        else:
            print("Skipped '{0}' - no given first name".format(cv))
    return cv_dicts


def get_slug(cv_dict):
    """
    emulates the cv model get_slug function with the dict data
    """
    return slugify(u"{0}-{1}-{2}".format(cv_dict["Given First Name"],
                                         cv_dict.get("Given Middle Name", ''),
                                         cv_dict.get("Surname", '')))

# synonyms map a synonym to a canonical value.
synonyms = {'Services On and Off-site': 'Services On/Off-site',
            'Given Name': 'Given First Name',
            'Given Second Name': 'Given Middle Name',
            'Date of Birth (mm/dd/yyyy)': 'Date of Birth',
            'Date of Birth (mm/dd/yy)': 'Date of Birth',
            'Family Name / Surname': 'Surname',
            '=HYPERLINK("http://en.wikipedia.org/wiki/ISO_3166-1", "Country")': 'Country',
            'UK (British)': 'United Kingdom',
            'French': 'France'}


def canonical(s):
    """
    Headings are not consistent across all CVS. Use canonical to eliminae the synonyms.
    """
    return synonyms.get(s, s)