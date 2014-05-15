"""
Load in initial data using the source excel spreadsheets
"""
from openpyxl import load_workbook

import cvs.models as cm
from cvs.importers.globals import *


def load_all():
    load_countries('.country-calc, CCCS')
    load_categorizations(get_theme_dict('.theme-calc, CCCS'), cm.CCCSTheme, cm.CCCSSubTheme, 'theme')
    load_categorizations(get_theme_dict('.theme-calc, IFC'), cm.IFCTheme, cm.IFCSubTheme, 'theme')
    load_ifc_sectors('.sector-calc, IFC')
    load_categorizations(get_cccs_sector_dict('.sector-calc, CCCS'), cm.CCCSSector, cm.CCCSSubSector, 'sector')
    load_projects('PROJECT')


def load_countries(sheet_name):
    country_dicts = get_country_dicts(sheet_name)
    for info in country_dicts:
        country, _ = cm.Country.objects.get_or_create(name=info['name'])
        for (k, v) in info.iteritems():
            setattr(country, k, v)
        country.save()


def get_country_dicts(sheet_name):
    """
    Use the gold cv to obtain the list of countries and related data
    Go for the '.country-calc, CCCS' sheet and collect the data from row 4 to the end.
    """
    wb = load_workbook(filename=xlsx_cv)
    ws = wb.get_sheet_by_name(sheet_name)

    countries = list()
    for row in ws.rows[3:]:  # ignore first three rows
        if row[0].value is None:
            continue  # ignore null entries
        countries.append({
            'name': row[0].value,
            'iso_english_name': row[1].value,
            'fips': row[2].value,
            'iso_numeric': row[3].value,
            'iso_3166': row[4].value,
            'iso': row[5].value,
            'notes': row[6].value})
    return countries


def load_categorizations(info, theme_model, sub_theme_model, super_field_name):
    for (theme_name, sub_theme_names) in info.iteritems():
        for sub_theme_name in sub_theme_names:
            _load_categorizations(theme_model, theme_name, sub_theme_model, sub_theme_name, super_field_name)


def get_theme_dict(sheet_name):
    """
    Use the gold cv to obtain all the cccs themes and sub themes
    """
    wb = load_workbook(filename=xlsx_cv)
    ws = wb.get_sheet_by_name(sheet_name)

    themes = dict()
    for row in ws.rows[1:]:  # ignore heading row
        theme_name = row[3].value
        sub_theme_name = row[4].value
        if theme_name is None:
            continue
        if theme_name not in themes:
            themes[theme_name] = set()
        themes[theme_name].add(sub_theme_name)
    return themes


def load_ifc_sectors(sheet_name):
    sector_names = get_ifc_sector_names(sheet_name)
    for sector_name in sector_names:
        sector, created = cm.IFCSector.objects.get_or_create(name=sector_name)
        if created:
            sector.save()


def get_ifc_sector_names(sheet_name):
    """
    Use the gold cv to obtain all the IFC sectors
    """
    wb = load_workbook(filename=xlsx_cv)
    ws = wb.get_sheet_by_name(sheet_name)
    sector_names = list()
    for row in ws.rows[2:]:  # ignore first two rows
        sector_name = row[1].value
        if sector_name is None:
            continue
        sector_names.append(sector_name)
    return sector_names


def get_cccs_sector_dict(sheet_name):
    """
    Use the gold cv to obtain all the cccs themes and sub themes
    """
    wb = load_workbook(filename=xlsx_cv)
    ws = wb.get_sheet_by_name(sheet_name)

    sectors = dict()
    for row in ws.rows[1:]:  # ignore heading row
        sector_name = row[1].value
        sub_sector_name = row[2].value
        if sector_name is None:
            continue
        if sector_name not in sectors:
            sectors[sector_name] = set()
        sectors[sector_name].add(sub_sector_name)
    return sectors


def load_projects(sheet_name):
    projects = get_project_dict(sheet_name)
    for project_name, project_info in projects.iteritems():
        project, _ = cm.Project.objects.get_or_create(name=project_name)
        project.name = project_name
        project.date_range = project_info['Date Range']
        project.loan_or_grant = project_info['Loan/TA/Grant No']
        project.features = project_info['Main Project Features']
        # Multiple countries are allowed but only one from import
        try:
            country = cm.Country.objects.get(name=project_info['Country'])
            project.countries.add(country)
        except cm.Country.DoesNotExist:
            print(u"No country for \"{0}\": '{1}' does not exist".format(project_name, project_info['Country']))
        project.region = project_info['Region']
        # Break Services On/Off-site into flags and have a go at importing them
        services = project_info.get('Services On/Off-site', None)
        if services is not None:
            project.service_on_site = 'On' in services
            project.service_off_site = 'Off' in services
            project.service_on_site = 'remote' in services
        project.position = project_info['Position']
        project.client_end = project_info['Sponsor / End Client']
        project.client_contract = project_info['Contracted Through / Direct Client']
        project.client_beneficiary = project_info['Beneficiary Client']
        project.contract = project_info['Contract No.']
        project.person_months = project_info['Person-months']
        project.activities = project_info['Activities Performed']
        project.references = project_info['References']
        for attname, args in (
                ('cccs_subthemes', (cm.CCCSTheme, project_info['Thematic Issues -GENERAL'],
                                    cm.CCCSSubTheme, project_info['Sub-Themes -GENERAL'],
                                    'theme')),
                ('cccs_subsectors', (cm.CCCSSector, project_info['Sector -GENERAL'],
                                     cm.CCCSSubSector, project_info['Sub-sector -GENERAL'],
                                     'sector')),
                ('ifc_subthemes', (cm.IFCTheme, project_info['Thematic Issues -IFC'],
                                   cm.IFCSubTheme, project_info['Sub-Themes -IFC'],
                                   'theme', False))):
            for categorization in _load_categorizations(*args):
                getattr(project, attname).add(categorization)

        ifc_sector_names = get_names_from_string(project_info['Sector -IFC'])
        for ifc_sector_name in ifc_sector_names:
            ifc_sector, created = cm.IFCSector.objects.get_or_create(name=ifc_sector_name)
            if created:
                ifc_sector.save()
            project.ifc_sectors.add(ifc_sector)

        project.save()


def _load_categorizations(super_class, super_values, sub_class, sub_values, super_field, add_unspecified=True):
    """
    Get or create the categorizations returning sub instances
    """
    if super_values:
        super_names = get_names_from_string(super_values)
    else:
        super_names = []
    if sub_values:
        sub_names = get_names_from_string(sub_values)
    else:
        sub_names = []

    if len(super_names) != len(sub_names):
        print("Mismatched theme/subtheme list lengths in cv - using shortest")
    if add_unspecified and not(super_names and sub_names):
        super_names = sub_names = ['Unspecified']

    sub_objects = list()
    for super_name, sub_name in zip(super_names, sub_names):
        super_obj, created = super_class.objects.get_or_create(name=super_name)
        if created:
            super_obj.save()

        sub_kwargs = {super_field: super_obj, 'name': sub_name}
        sub_obj, created = sub_class.objects.get_or_create(**sub_kwargs)
        if created:
            sub_obj.save()
        sub_objects.append(sub_obj)

    return sub_objects


def get_names_from_string(s):
    if s is None:
        return []
    else:
        return [name.strip() for name in s.split(';') if name]


def get_project_dict(sheet_name):
    projects = dict()
    for cv in xlsx_cvs:
        try:
            wb = load_workbook(filename=cv)
        except TypeError:
            print("Unable to open {0}".format(cv))
            continue
        ws = wb.get_sheet_by_name(sheet_name)
        headings = _get_project_dict_headings(ws.rows[0])
        for row in ws.rows[1:]:
            project_name = row[1].value
            if project_name is not None:
                if project_name in projects:
                    # Add any values we don't already have
                    existing_info = projects[project_name]
                    for k, c in zip(headings, row):
                        new_value = c.value
                        if new_value is None:
                            continue
                        if existing_info[k] is None:
                            existing_info[k] = new_value
                        elif new_value != existing_info[k]:
                            print(u"Mismatched project info: '{0}' vs '{1}'".format(new_value, existing_info[k]))
                            # Use longest or biggest
                            try:
                                if len(new_value) > len(existing_info[k]):
                                    existing_info[k] = new_value
                            except TypeError:
                                try:
                                    if new_value > existing_info[k]:
                                        existing_info[k] = new_value
                                except TypeError:
                                    pass
                else:
                    # Add initial values
                    projects[project_name] = {k: c.value for (k, c) in zip(headings, row)}
    return projects


def _get_project_dict_headings(heading_row):
    headings = list()
    for heading_cell in heading_row:
        if heading_cell.value is None:
            break
        headings.append(canonical(heading_cell.value.strip()))
    return headings

# synonyms map a synonym to a canonical value.
synonyms = {'Services On and Off-site': "Services On/Off-site"}


def canonical(s):
    """
    Headings are not consistent across all CVS. Here we define canonical forms so that the synonyms
    don't screw up the importing process.
    (Lower all case to reduce likely errors)
    """
    return synonyms.get(s, s)