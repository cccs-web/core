"""
Load in initial data using the source excel spreadsheets
"""
import os

from openpyxl import load_workbook

import cvs.models as cm

pwd = os.path.dirname(__file__)
xlsx_cvs = [os.path.abspath(os.path.join(pwd, 'source_data', fn))
            for fn in os.listdir(os.path.join(pwd, 'source_data'))
            if fn.startswith('CV')]

xlsx_cv = next((cv for cv in xlsx_cvs if 'Aaron' in cv))


def load_all():
    load_countries('.country-calc, CCCS')
    load_themes('.theme-calc, CCCS', cm.CCCSTheme, cm.CCCSSubTheme)
    load_themes('.theme-calc, IFC', cm.IFCTheme, cm.IFCSubTheme)
    load_sectors('.sector-calc, IFC')


def load_countries(sheet_name):
    country_dicts = get_country_dicts(sheet_name)
    for info in country_dicts:
        country, _ = cm.Country.objects.get_or_create(name=info['name'])
        for (k, v) in info.iteritems():
            setattr(country, k, v)
        country.save()


def get_country_dicts(sheet_name):
    """
    Use the gold cv to obtain the list of countries and related data
    Go for the '.country-calc, CCCS' sheet and collect the data from row 4 to the end.
    """
    wb = load_workbook(filename=xlsx_cv)
    ws = wb.get_sheet_by_name(sheet_name)

    countries = list()
    for row in ws.rows[3:]:  # ignore first three rows
        if row[0].value is None:
            continue  # ignore null entries
        countries.append({
            'name': row[0].value,
            'iso_english_name': row[1].value,
            'fips': row[2].value,
            'iso_numeric': row[3].value,
            'iso_3166': row[4].value,
            'iso': row[5].value,
            'notes': row[6].value})
    return countries


def load_themes(sheet_name, theme_model, sub_theme_model):
    theme_dict = get_theme_dict(sheet_name)
    for (theme_name, sub_theme_names) in theme_dict.iteritems():
        theme, created = theme_model.objects.get_or_create(name=theme_name)
        if created:
            theme.save()
        for sub_theme_name in sub_theme_names:
            sub_theme, created = sub_theme_model.objects.get_or_create(theme=theme, name=sub_theme_name)
            if created:
                sub_theme.save()


def get_theme_dict(sheet_name):
    """
    Use the gold cv to obtain all the cccs themes and sub themes
    """
    wb = load_workbook(filename=xlsx_cv)
    ws = wb.get_sheet_by_name(sheet_name)

    themes = dict()
    for row in ws.rows[1:]:  # ignore heading row
        theme_name = row[3].value
        sub_theme_name = row[4].value
        if theme_name is None:
            continue
        if theme_name not in themes:
            themes[theme_name] = set()
        themes[theme_name].add(sub_theme_name)
    return themes


def load_sectors(sheet_name):
    sector_names = get_sector_names(sheet_name)
    for sector_name in sector_names:
        sector, created = cm.IFCSector.objects.get_or_create(name=sector_name)
        if created:
            sector.save()


def get_sector_names(sheet_name):
    """
    Use the gold cv to obtain all the IFC sectors
    """
    wb = load_workbook(filename=xlsx_cv)
    ws = wb.get_sheet_by_name(sheet_name)
    sector_names = list()
    for row in ws.rows[2:]:  # ignore first two rows
        sector_name = row[1].value
        if sector_name is None:
            continue
        sector_names.append(sector_name)
    return sector_names